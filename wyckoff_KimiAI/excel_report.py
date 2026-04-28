"""
Excel报告生成模块
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from datetime import datetime
from typing import List, Dict
import logging

logger = logging.getLogger('ExcelReport')

class ExcelReportGenerator:
    """Excel报告生成器"""
    
    def __init__(self, template_config: Dict = None):
        self.template = template_config or {
            'header_color': '4472C4',
            'header_font_color': 'FFFFFF',
            'excellent_color': 'C6EFCE',  # 浅绿 80+
            'good_color': 'FFEB9C',        # 浅黄 70-79
            'normal_color': 'FFFFFF',      # 白色 60-69
            'warning_color': 'FFC7CE'      # 浅红 <60
        }
    
    def generate_report(self, results: List, data_manager, 
                       output_path: Path = None) -> Path:
        """
        生成完整Excel报告
        """
        if output_path is None:
            output_path = Path(f"./reports/wyckoff_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        output_path.parent.mkdir(exist_ok=True)
        
        # 创建工作簿
        wb = Workbook()
        
        # 1. 摘要页
        self._create_summary_sheet(wb, results)
        
        # 2. 详细结果页
        self._create_detail_sheet(wb, results)
        
        # 3. 行业分析页
        self._create_industry_sheet(wb, results)
        
        # 4. 阶段分布页
        self._create_phase_sheet(wb, results)
        
        # 5. 个股详情页（前20只）
        self._create_stock_details(wb, results[:20], data_manager)
        
        # 保存
        wb.save(output_path)
        logger.info(f"Excel报告已生成: {output_path}")
        
        return output_path
    
    def _create_summary_sheet(self, wb: Workbook, results: List):
        """创建摘要页"""
        ws = wb.active
        ws.title = "摘要"
        
        # 标题
        ws['A1'] = "威科夫选股系统 - 分析报告"
        ws['A1'].font = Font(size=16, bold=True, color=self.template['header_color'])
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 生成时间
        ws['A3'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = Font(size=10, italic=True)
        
        # 统计摘要
        stats = self._calculate_stats(results)
        
        row = 5
        headers = ['统计项', '数值']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color=self.template['header_font_color'])
            cell.fill = PatternFill(start_color=self.template['header_color'], end_color=self.template['header_color'], fill_type='solid')
        
        row += 1
        for key, value in stats.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def _create_detail_sheet(self, wb: Workbook, results: List):
        """创建详细结果页"""
        ws = wb.create_sheet("详细结果")
        
        # 转换为DataFrame
        if results and hasattr(results[0], 'symbol'):
            # StockAnalysis对象列表
            df_data = []
            for r in results:
                df_data.append({
                    '排名': 0,  # 稍后填充
                    '代码': r.symbol,
                    '名称': r.name,
                    '行业': r.industry,
                    '威科夫阶段': r.phase.value,
                    '评分': r.composite_score,
                    '当前价': r.technicals.get('current_price', 0),
                    '支撑位': r.technicals.get('support', 0),
                    '阻力位': r.technicals.get('resistance', 0),
                    'RS(%)': round(r.technicals.get('rs', 0), 2),
                    '建议': r.recommendation,
                    '最新信号': r.signals[0].event.value if r.signals else '无',
                    '数据新鲜度': r.data_freshness
                })
            df = pd.DataFrame(df_data)
            df['排名'] = range(1, len(df) + 1)
        else:
            # 已经是DataFrame或字典列表
            df = pd.DataFrame(results)
            if '排名' not in df.columns:
                df.insert(0, '排名', range(1, len(df) + 1))
        
        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # 表头样式
                if r_idx == 1:
                    cell.font = Font(bold=True, color=self.template['header_font_color'])
                    cell.fill = PatternFill(start_color=self.template['header_color'], 
                                           end_color=self.template['header_color'], fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                else:
                    # 数据行颜色（根据评分）
                    score_col = None
                    for idx, col_name in enumerate(df.columns, 1):
                        if col_name == '评分':
                            score_col = idx
                            break
                    
                    if score_col and c_idx == score_col:
                        score = value if isinstance(value, (int, float)) else 0
                        if score >= 80:
                            cell.fill = PatternFill(start_color=self.template['excellent_color'], 
                                                   end_color=self.template['excellent_color'], fill_type='solid')
                        elif score >= 70:
                            cell.fill = PatternFill(start_color=self.template['good_color'], 
                                                   end_color=self.template['good_color'], fill_type='solid')
                        elif score < 60:
                            cell.fill = PatternFill(start_color=self.template['warning_color'], 
                                                   end_color=self.template['warning_color'], fill_type='solid')
        
        # 设置列宽
        for idx, col in enumerate(df.columns, 1):
            ws.column_dimensions[chr(64 + idx) if idx <= 26 else 'A' + chr(64 + idx - 26)].width = 12
        
        # 添加图表：评分分布
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "威科夫评分分布"
        chart.y_axis.title = '股票数量'
        chart.x_axis.title = '评分区间'
        
        # 统计各分数段数量
        score_bins = pd.cut(df['评分'], bins=[0, 60, 70, 80, 100], labels=['<60', '60-69', '70-79', '80+'])
        score_counts = score_bins.value_counts().sort_index()
        
        # 写入图表数据
        chart_row = len(df) + 5
        ws.cell(row=chart_row, column=1, value="评分区间")
        ws.cell(row=chart_row, column=2, value="数量")
        for i, (interval, count) in enumerate(score_counts.items(), 1):
            ws.cell(row=chart_row + i, column=1, value=str(interval))
            ws.cell(row=chart_row + i, column=2, value=count)
        
        data = Reference(ws, min_col=2, min_row=chart_row, max_row=chart_row + len(score_counts))
        cats = Reference(ws, min_col=1, min_row=chart_row + 1, max_row=chart_row + len(score_counts))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws.add_chart(chart, f"D{chart_row}")
    
    def _create_industry_sheet(self, wb: Workbook, results: List):
        """创建行业分析页"""
        ws = wb.create_sheet("行业分析")
        
        # 统计行业分布
        if results and hasattr(results[0], 'industry'):
            industries = {}
            for r in results:
                ind = r.industry
                if ind not in industries:
                    industries[ind] = {'count': 0, 'avg_score': 0, 'total_score': 0}
                industries[ind]['count'] += 1
                industries[ind]['total_score'] += r.composite_score
            
            for ind in industries:
                industries[ind]['avg_score'] = industries[ind]['total_score'] / industries[ind]['count']
            
            # 写入
            headers = ['行业', '股票数量', '平均评分', '占比(%)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color=self.template['header_font_color'])
                cell.fill = PatternFill(start_color=self.template['header_color'], 
                                       end_color=self.template['header_color'], fill_type='solid')
            
            total = len(results)
            sorted_industries = sorted(industries.items(), key=lambda x: x[1]['count'], reverse=True)
            
            for row, (ind, data) in enumerate(sorted_industries, 2):
                ws.cell(row=row, column=1, value=ind)
                ws.cell(row=row, column=2, value=data['count'])
                ws.cell(row=row, column=3, value=round(data['avg_score'], 2))
                ws.cell(row=row, column=4, value=round(data['count'] / total * 100, 2))
        
        # 设置列宽
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 15
    
    def _create_phase_sheet(self, wb: Workbook, results: List):
        """创建阶段分布页"""
        ws = wb.create_sheet("阶段分布")
        
        # 统计阶段分布
        if results and hasattr(results[0], 'phase'):
            phases = {}
            for r in results:
                phase = r.phase.value
                if phase not in phases:
                    phases[phase] = {'count': 0, 'avg_score': 0}
                phases[phase]['count'] += 1
                phases[phase]['avg_score'] += r.composite_score
            
            for phase in phases:
                phases[phase]['avg_score'] /= phases[phase]['count']
            
            # 写入
            headers = ['威科夫阶段', '股票数量', '平均评分', '占比(%)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color=self.template['header_font_color'])
                cell.fill = PatternFill(start_color=self.template['header_color'], 
                                       end_color=self.template['header_color'], fill_type='solid')
            
            total = len(results)
            row = 2
            for phase, data in phases.items():
                ws.cell(row=row, column=1, value=phase)
                ws.cell(row=row, column=2, value=data['count'])
                ws.cell(row=row, column=3, value=round(data['avg_score'], 2))
                ws.cell(row=row, column=4, value=round(data['count'] / total * 100, 2))
                row += 1
        
        # 设置列宽
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 15
    
    def _create_stock_details(self, wb: Workbook, results: List, data_manager):
        """创建个股详情页（前20只）"""
        for result in results:
            try:
                symbol = result.symbol if hasattr(result, 'symbol') else result['代码']
                name = result.name if hasattr(result, 'name') else result['名称']
                
                ws = wb.create_sheet(f"{symbol}_{name[:4]}")
                
                # 基本信息
                ws['A1'] = f"{symbol} {name} - 详细分析"
                ws['A1'].font = Font(size=14, bold=True)
                ws.merge_cells('A1:D1')
                
                # 技术指标
                if hasattr(result, 'technicals'):
                    tech = result.technicals
                    row = 3
                    ws.cell(row=row, column=1, value="技术指标")
                    ws.cell(row=row, column=1).font = Font(bold=True)
                    
                    metrics = [
                        ('当前价格', f"{tech.get('current_price', 0):.2f}"),
                        ('20日均线', f"{tech.get('ma20', 0):.2f}"),
                        ('60日均线', f"{tech.get('ma60', 0):.2f}"),
                        ('支撑位', f"{tech.get('support', 0):.2f}"),
                        ('阻力位', f"{tech.get('resistance', 0):.2f}"),
                        ('RS(%)', f"{tech.get('rs', 0):.2f}"),
                        ('ATR(%)', f"{tech.get('atr_pct', 0):.2f}"),
                        ('量比', f"{tech.get('volume_ratio', 0):.2f}"),
                    ]
                    
                    for i, (metric, value) in enumerate(metrics, 4):
                        ws.cell(row=i, column=1, value=metric)
                        ws.cell(row=i, column=2, value=value)
                
                # 信号记录
                if hasattr(result, 'signals') and result.signals:
                    row = 13
                    ws.cell(row=row, column=1, value="威科夫信号记录")
                    ws.cell(row=row, column=1).font = Font(bold=True)
                    
                    headers = ['日期', '信号类型', '置信度', '描述', '价格']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row+1, column=col, value=header)
                        cell.font = Font(bold=True, color=self.template['header_font_color'])
                        cell.fill = PatternFill(start_color=self.template['header_color'], 
                                               end_color=self.template['header_color'], fill_type='solid')
                    
                    for i, signal in enumerate(result.signals, row+2):
                        ws.cell(row=i, column=1, value=signal.date)
                        ws.cell(row=i, column=2, value=signal.event.value)
                        ws.cell(row=i, column=3, value=f"{signal.confidence:.0%}")
                        ws.cell(row=i, column=4, value=signal.description)
                        ws.cell(row=i, column=5, value=signal.price)
                
                # 设置列宽
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 40
                ws.column_dimensions['E'].width = 12
                
            except Exception as e:
                logger.error(f"创建 {symbol} 详情页失败: {e}")
                continue
    
    def _calculate_stats(self, results: List) -> Dict:
        """计算统计摘要"""
        if not results:
            return {'总股票数': 0}
        
        if hasattr(results[0], 'composite_score'):
            scores = [r.composite_score for r in results]
            phases = [r.phase.value for r in results]
        else:
            scores = [r.get('评分', 0) for r in results]
            phases = [r.get('威科夫阶段', '未知') for r in results]
        
        return {
            '总股票数': len(results),
            '平均评分': round(sum(scores) / len(scores), 2),
            '最高评分': max(scores),
            '最低评分': min(scores),
            '80分以上': len([s for s in scores if s >= 80]),
            '70-79分': len([s for s in scores if 70 <= s < 80]),
            '60-69分': len([s for s in scores if 60 <= s < 70]),
            '吸筹阶段': phases.count('吸筹阶段'),
            '上涨阶段': phases.count('上涨阶段'),
            '派发阶段': phases.count('派发阶段'),
        }
